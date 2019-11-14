import openpyxl

# wb = openpyxl.workbook()
wb = openpyxl.load_workbook('Marketing Digital_Estratégias_Negócios.xlsx')
print(wb.sheetnames)

sheet = wb['Concorrentes']

cell = sheet['a1']
"""
Methods to Access cells:
cell = sheet['a1']
column = sheet['a']
cells = sheet['a:c']
sheet['a1:c3']
sheet[1:3]
"""
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)
sheet.append([1, 2, 3])
wb.save('Marketing2.xlsx')


